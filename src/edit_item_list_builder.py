"""
edit_item_list_builder.py

자동화 영역: qoo10_item_detail_scraper.py로 수집한 상품 정보 +
category_brand_matcher.py로 조회한 카테고리/브랜드 코드를 합쳐서,
공식 Qoo10_EditItemList.xlsx 템플릿의 헤더/가이드 행(1~4행)은 그대로 두고
5행부터 데이터 행을 채운 새 업로드 파일을 만든다.

검수 게이트: match_review_builder.py가 만든 결정 파일(decisions.json)을 넘기면,
match_confirmed=true 인 상품만 업로드 양식에 포함된다. 결정 파일을 넘기지 않으면
검수 없이 전부 포함하므로 사람이 눈으로 확인하지 않은 상품이 그대로 올라갈 수
있다 — 실제 업로드 전에는 반드시 결정 파일을 함께 넘길 것.
image_usable=true 로 확인되지 않은 상품은 결정 파일이 있어도 image_main_url이
자동으로 TODO 처리된다(이미지 저작권 확인 전에는 채우지 않음).

[상품명 규칙] item_name은 항상 큐텐 원본 상세페이지의 원문 그대로 채운다.
검색 단계에서 쓰는 핵심문구(브랜드+고유명+용량 축약형)나, 매칭 검수용으로
찾은 한글 상품명(match_review_builder.py의 name_kr)은 검색/확인 용도일 뿐이며
업로드 필드에는 절대 섞이지 않는다:
    큐텐 원본 상품명 --(그대로)--> item_name(업로드용)
    큐텐 원본 상품명 --(축약)--> 핵심문구 --(검색)--> 한글 상품명(사람 확인용, 업로드 미사용)

자동으로 채우는 필드 (신뢰도 높음):
    seller_unique_item_id, item_name, image_main_url(검수 승인 시, 고화질 URL 우선),
    category_number(큐텐 상품페이지에서 원 판매자가 지정한 소카테고리 코드를 그대로 재사용),
    brand_number(단일 후보일 때만),
    item_status_Y/N/D, end_date, quantity, Shipping_number,
    available_shipping_date, item_condition_type, origin_type, origin_country_id,
    price_yen(korea_side.json에 price_krw + 카테고리 무게참고값이 모두 있을 때
              margin_calculator.py로 실제 산식 계산 — Q10_계산기.xlsx "마진율 먼저 설정" 방식 재현)

TODO로 남기는 필드 (AI/사람 판단 필요, README 참고):
    price_yen                     -> korea_side.json의 원가(price_krw) 또는 카테고리
                                      무게참고값이 없을 때만
    retail_price_yen              -> 정가(할인 연출용) 표시는 계산기 범위 밖, 비즈니스 판단 필요
    category_number               -> 큐텐 페이지에서 코드 추출 실패했을 때만 (드묾)
    brand_number                  -> 매칭 후보가 여러 개거나 없을 때
    image_main_url                -> image_usable=true 로 검수 승인되지 않았을 때
    image_other_url, item_description -> 상세페이지 구조가 셀러마다 달라 미추출
    item_weight                   -> 실측값 아님. 카테고리 실제이력 참고표(있으면) 또는
                                      상품명 추정치를 메모로만 표시, 최종 확정은 사람이

사용법:
    python edit_item_list_builder.py <template.xlsx> <items_dir> <output.xlsx> \
        [<decisions.json>] [<korea_side.json>] [목표마진율=0.12] [환율(원/100엔)=900]

    items_dir 안의 각 <goods_no>.json 은 qoo10_item_detail_scraper.py 출력 형식이어야 한다.
    decisions.json 은 match_review_builder.py 출력 형식이어야 한다(생략 가능하나 비권장).
    korea_side.json 은 match_review_builder.py가 참조하는 것과 동일한 형식으로,
    각 항목의 price_krw가 "국내구매가"로 쓰여 마진계산기에 들어간다(생략 시 price_yen은 TODO).
"""

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from category_brand_matcher import BrandCategoryMatcher
import margin_calculator

DATA_START_ROW = 5  # 템플릿의 예시 행부터 실제 데이터로 채움
TODO = "TODO"

# 신규 등록 시 공통 기본값 (사용자의 실제 등록 이력 322건 전부 동일하게 확인된 값 —
# Qoo10_ItemInfo_20260719204313.xlsx 실증 분석 결과로 교체함)
DEFAULTS = {
    "item_status_Y/N/D": "Y",
    "end_date": "2050-12-31 00:00:00",
    "quantity": 200,
    "Shipping_number": 741315,  # 실제 등록이력 322건 전부 이 코드 사용 (기존 0은 템플릿 예시값이었을 뿐)
    "available_shipping_date": 3,
    "item_condition_type": 1,  # 1: 새상품
    "origin_type": 2,  # 2: 해외
    "origin_country_id": "KR",  # 한국 소싱 상품 기준
}


def load_weight_reference() -> dict:
    """실제 등록이력에서 뽑은 category_number -> 무게(kg) 참고표를 로드한다."""
    data_dir = Path(__file__).resolve().parent.parent / "data"
    path = data_dir / "weight_by_category.json"
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def load_items(items_dir: str) -> list[dict]:
    items = []
    for path in sorted(Path(items_dir).glob("*.json")):
        items.append(json.loads(path.read_text(encoding="utf-8")))
    return items


def load_decisions(decisions_path: str | None) -> dict:
    """match_review_builder.py가 만든 결정 파일을 goods_no 기준 dict로 로드한다."""
    if not decisions_path:
        return {}
    decisions = json.loads(Path(decisions_path).read_text(encoding="utf-8"))
    return {d["goods_no"]: d for d in decisions if d.get("goods_no")}


def load_korea_side(korea_side_path: str | None) -> dict:
    """korea_side.json(사람이 채운 한국 쪽 가격정보)을 goods_no 기준 dict로 로드한다.
    price_krw가 "국내구매가"로 쓰여 마진계산기 입력값이 된다."""
    if not korea_side_path:
        return {}
    data = json.loads(Path(korea_side_path).read_text(encoding="utf-8"))
    return {d["goods_no"]: d for d in data if d.get("goods_no")}


def build_row(
    item: dict,
    matcher: BrandCategoryMatcher,
    decision: dict | None = None,
    weight_ref: dict | None = None,
    korea_side_item: dict | None = None,
    margin_rate: float = margin_calculator.DEFAULT_MARGIN_RATE,
    exchange_rate: float = 900,
) -> dict:
    row = dict(DEFAULTS)
    weight_ref = weight_ref or {}

    row["seller_unique_item_id"] = item.get("goods_no", TODO)

    # [중요] item_name은 반드시 큐텐 원본 상세페이지에서 스크랩한 원문 그대로 사용한다.
    # 검색용으로 만든 핵심문구(브랜드+고유명+용량 축약형)나 매칭 확인용 한글 상품명(name_kr)은
    # 절대 여기 들어가면 안 된다 — 이 함수는 qoo10_item_detail_scraper.py가 넘긴 item dict의
    # "item_name" 키(큐텐 JSON-LD의 원본 상품명)만 참조한다.
    qoo10_original_name = item.get("item_name")
    assert not isinstance(decision, dict) or "name_kr" not in row, "item_name에 한글명이 섞이면 안 됨"
    row["item_name"] = qoo10_original_name or TODO

    image_usable = (decision or {}).get("image_usable")
    if image_usable is True:
        # 검수 페이지에서 사람이 직접 고른 사진(final_image)을 최우선으로 쓴다.
        # (큐텐 사진 2장/한국 사진 2장 중 선택한 결과 — match_review_builder.py 참고)
        chosen = (decision or {}).get("final_image")
        row["image_main_url"] = chosen or item.get("image_main_url_hires") or item.get("image_main_url") or TODO
    else:
        row["image_main_url"] = (
            f"{TODO} (이미지 사용 승인 안됨 — match_review_builder.py 결정 파일에서 "
            "image_usable을 true로 확인 후 재실행하거나, 직접 촬영/구매한 이미지로 교체)"
        )

    brand_name = item.get("brand_name")
    if brand_name:
        candidates = matcher.find_brand(brand_name)
        if len(candidates) == 1:
            row["brand_number"] = candidates[0]["brand_no"]
        else:
            row["brand_number"] = f"{TODO} (후보 {len(candidates)}건, brand_name={brand_name})"
    else:
        row["brand_number"] = TODO

    # 큐텐 상품페이지에 원 판매자가 지정해둔 소카테고리 코드를 그대로 재사용한다.
    # (Qoo10_CategoryInfo.csv의 소카테고리 코드와 동일한 체계) 단, 이건 "원 판매자"가
    # 붙인 분류라 우리 상품 맥락과 다를 수 있으므로 최종 확인은 권장 — TODO는 아니지만
    # 사람이 한 번 훑어보는 걸 권장한다는 뜻에서 category_gdsc_cd가 없을 때만 TODO.
    gdsc_cd = item.get("category_gdsc_cd")
    row["category_number"] = gdsc_cd if gdsc_cd else f"{TODO} (item_name 참고해서 카테고리 지정 필요)"

    cat_weight = weight_ref.get(gdsc_cd) if gdsc_cd else None
    cost_krw = (korea_side_item or {}).get("price_krw")

    if cost_krw and cat_weight:
        # 마진계산기(Q10_계산기.xlsx "2.마진율 먼저 설정" 방식)로 실제 판매가를 역산한다.
        calc = margin_calculator.calculate(
            cost_krw=cost_krw,
            weight_kg=cat_weight["median_kg"],
            margin_rate=margin_rate,
            exchange_rate=exchange_rate,
        )
        row["price_yen"] = calc["price_yen"]
        row["retail_price_yen"] = TODO  # 정가(할인 연출용)는 계산기 범위 밖 — 비즈니스 판단 필요
    else:
        missing = []
        if not cost_krw:
            missing.append("한국 국내구매가(korea_side.json의 price_krw)")
        if not cat_weight:
            missing.append("카테고리 무게 참고값")
        row["price_yen"] = (
            f"{TODO} (마진계산기 계산 불가 — {', '.join(missing)} 없음. "
            f"큐텐참고가={item.get('price_jpy')}円)"
        )
        row["retail_price_yen"] = TODO

    row["image_other_url"] = TODO
    row["item_description"] = TODO

    weight_hint = item.get("weight_hint")
    if cat_weight:
        row["item_weight"] = (
            f"{TODO} (참고: 동일 카테고리 실제 등록이력 {cat_weight['sample_count']}건 기준 "
            f"중간값={cat_weight['median_kg']}kg, 범위={cat_weight['min_kg']}~{cat_weight['max_kg']}kg"
            + (f" / 상품명 추정={weight_hint}" if weight_hint else "")
            + " — 실측 아님, 확인 후 기입)"
        )
    elif weight_hint:
        row["item_weight"] = f"{TODO} (참고: 상품명에서 추정한 값={weight_hint}, 실측 아님 — 확인 후 기입)"
    else:
        row["item_weight"] = TODO

    return row


COLUMN_ORDER = [
    "item_number", "seller_unique_item_id", "category_number", "brand_number", "item_name",
    "item_promotion_name", "item_status_Y/N/D", "start_date", "end_date", "price_yen",
    "retail_price_yen", "taxrate", "quantity", "option_info", "additional_option_info",
    "additional_option_text", "image_main_url", "image_other_url", "video_url",
    "image_option_info", "image_additional_option_info", "header_html", "footer_html",
    "item_description", "Shipping_number", "option_number", "available_shipping_date",
    "desired_shipping_date", "search_keyword", "item_condition_type", "origin_type",
    "origin_region_id", "origin_country_id", "origin_others", "medication_type",
    "item_weight", "item_material", "model_name", "external_product_type",
    "external_product_id", "manufacture_date", "expiration_date_type", "expiration_date_MFD",
    "expiration_date_PAO", "expiration_date_EXP", "under18s_display_Y/N", "A/S_info",
    "buy_limit_type", "buy_limit_date", "buy_limit_qty",
]


def build_workbook(
    template_path: str,
    items: list[dict],
    matcher: BrandCategoryMatcher,
    out_path: str,
    decisions: dict | None = None,
    weight_ref: dict | None = None,
    korea_side: dict | None = None,
    margin_rate: float = margin_calculator.DEFAULT_MARGIN_RATE,
    exchange_rate: float = 900,
) -> tuple[int, int]:
    """반환값: (작성된 행 수, 스킵된 행 수)"""
    decisions = decisions or {}
    weight_ref = weight_ref or {}
    korea_side = korea_side or {}
    wb = load_workbook(template_path)
    ws = wb.active

    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert header_row == COLUMN_ORDER, "템플릿 컬럼 순서가 예상과 다릅니다. 템플릿이 바뀌었는지 확인하세요."

    written = 0
    skipped = 0
    for item in items:
        goods_no = item.get("goods_no")
        decision = decisions.get(goods_no)

        if decisions and (not decision or decision.get("match_confirmed") is not True):
            print(
                f"[SKIP] goods_no={goods_no} — match_review_builder.py 결정 파일에서 "
                "match_confirmed=true 로 확정되지 않아 업로드 양식에서 제외합니다."
            )
            skipped += 1
            continue

        korea_side_item = korea_side.get(goods_no)
        row_data = build_row(item, matcher, decision, weight_ref, korea_side_item, margin_rate, exchange_rate)
        r = DATA_START_ROW + written
        for col_idx, col_name in enumerate(COLUMN_ORDER, start=1):
            if col_name in row_data:
                ws.cell(row=r, column=col_idx, value=row_data[col_name])
        written += 1

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return written, skipped


def main():
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)

    template_path, items_dir, out_path = sys.argv[1:4]
    decisions_path = sys.argv[4] if len(sys.argv) > 4 else None
    korea_side_path = sys.argv[5] if len(sys.argv) > 5 else None
    margin_rate = float(sys.argv[6]) if len(sys.argv) > 6 else margin_calculator.DEFAULT_MARGIN_RATE
    exchange_rate = float(sys.argv[7]) if len(sys.argv) > 7 else 900

    data_dir = Path(__file__).resolve().parent.parent / "data"
    matcher = BrandCategoryMatcher(
        str(data_dir / "brand_list.csv"), str(data_dir / "qoo10_category_info.csv")
    )

    items = load_items(items_dir)
    decisions = load_decisions(decisions_path)
    weight_ref = load_weight_reference()
    korea_side = load_korea_side(korea_side_path)
    print(f"[INFO] {len(items)}건 로드" + (f", 결정파일 {len(decisions)}건 로드" if decisions else " (결정파일 없음 — 검수 없이 전부 포함, 권장하지 않음)"))
    print(f"[INFO] 무게 참고표 {len(weight_ref)}개 카테고리 로드")
    print(f"[INFO] 한국측 원가정보 {len(korea_side)}건 로드" + (f", 목표마진율={margin_rate}, 환율={exchange_rate}" if korea_side else " (없음 — price_yen은 TODO로 남음)"))

    written, skipped = build_workbook(
        template_path, items, matcher, out_path, decisions, weight_ref, korea_side, margin_rate, exchange_rate
    )
    print(f"[INFO] 작성 완료 -> {out_path} ({written}건 작성, {skipped}건 검수 미승인으로 제외)")
    print("[INFO] TODO 표시된 셀(카테고리/설명 등)은 반드시 사람이 채운 뒤 업로드하세요.")


if __name__ == "__main__":
    main()
