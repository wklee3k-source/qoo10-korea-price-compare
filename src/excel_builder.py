"""
excel_builder.py

자동화 영역: 큐텐/한국 상품 비교 데이터(JSON)를 받아 이미지가 삽입된 엑셀로 출력한다.

입력 JSON 형식 (리스트, 각 항목):
    {
      "brand_ja": str, "name_ja": str, "price_jpy": int, "img_qoo10": str(local path),
      "name_kr": str, "price_krw": int, "img_kr": str(local path), "kr_site": str
    }

사용법:
    python excel_builder.py <input.json> <output.xlsx>
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "No", "브랜드", "큐텐 상품명(일본어)", "큐텐 사진", "큐텐 가격(円)",
    "한국 상품명", "한국 사진", "한국 가격(원)", "한국 판매처",
]

COL_WIDTHS = {1: 5, 2: 16, 3: 34, 4: 14, 5: 12, 6: 30, 7: 14, 8: 12, 9: 22}
IMG_SIZE = 90


def build_excel(items: list[dict], out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "큐텐-한국 상품비교"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[1].height = 22

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    row = 2
    for i, item in enumerate(items):
        ws.row_dimensions[row].height = 70

        ws.cell(row=row, column=1, value=i + 1).alignment = center
        ws.cell(row=row, column=2, value=item.get("brand_ja", "")).alignment = center
        ws.cell(row=row, column=3, value=item.get("name_ja", "")).alignment = left

        c5 = ws.cell(row=row, column=5, value=item.get("price_jpy"))
        c5.number_format = '#,##0"円"'
        c5.alignment = center

        ws.cell(row=row, column=6, value=item.get("name_kr", "")).alignment = left

        c8 = ws.cell(row=row, column=8, value=item.get("price_krw"))
        c8.number_format = '#,##0"원"'
        c8.alignment = center

        ws.cell(row=row, column=9, value=item.get("kr_site", "")).alignment = left

        for col in range(1, 10):
            ws.cell(row=row, column=col).border = border
            if col not in (5, 8):
                ws.cell(row=row, column=col).font = Font(name="Arial", size=10)

        img_qoo10 = item.get("img_qoo10")
        if img_qoo10 and Path(img_qoo10).exists():
            img1 = XLImage(img_qoo10)
            img1.width = IMG_SIZE
            img1.height = IMG_SIZE
            ws.add_image(img1, f"D{row}")

        img_kr = item.get("img_kr")
        if img_kr and Path(img_kr).exists():
            img2 = XLImage(img_kr)
            img2.width = IMG_SIZE
            img2.height = IMG_SIZE
            ws.add_image(img2, f"G{row}")

        row += 1

    ws.freeze_panes = "A2"
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    items = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    build_excel(items, sys.argv[2])
    print(f"[INFO] wrote {sys.argv[2]}")


if __name__ == "__main__":
    main()
