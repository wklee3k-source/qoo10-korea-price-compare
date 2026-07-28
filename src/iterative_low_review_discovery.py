"""
iterative_low_review_discovery.py

사용자가 정의한 알고리즘(v2, 명확화됨):

    1. (초기) 검색어로 큐텐 검색 → 리뷰 없음/3개 미만 상점 찾기
    2. 그 상점의 베스트5를 크롤링한다. 크롤링 시점에 카테고리가 색조
       (베이스메이크업/포인트메이크업/메이크업소품 — 반드시 옵션이 생기는
       계열)면 스킵, 옵션이 있으면 스킵. 통과한 상품은 원본 상품명 그대로 저장.
    3. 통과한 상품명에서 핵심단어를 추출한다(괄호/슬래시이후/수량단위/
       잡음단어 제거).
    4. 핵심단어로 재검색해서 리뷰 없음/3개 미만 상점을 또 찾는다.
    5. 그 상점들의 베스트5를 크롤링(2번부터 반복) — 상점은 전체 라운드에
       걸쳐 중복 방문 안 함.

사용법:
    python iterative_low_review_discovery.py "<초기검색어(일본어)>" <목표상품수> <output.xlsx> [최대상점수]
"""

import json
import re
import subprocess
import sys
from pathlib import Path

from qoo10_low_review_shop_finder import search_qoo10, parse_results
from qoo10_ranking_scraper import fetch_shop_ranking, ShopCrawlFailed
from qoo10_item_detail_scraper import fetch_item_detail
from google_translate import GoogleTranslateSession
from hwahae_name_corrector import correct_name

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"
STATE_PATH = OUTPUT_DIR / "discovery_state.json"

COLOR_COSMETIC_CATEGORIES = {"120000013", "120000014", "120000016"}

# 화장품/뷰티 허용 대분류 코드(화이트리스트) — 이게 없으면 속옷/식품/잡화 같은
# 완전히 무관한 카테고리(예: "흰비둘기 거들", "롯데 가나 초콜릿 쿠키")가
# 그대로 통과하는 문제가 실측으로 확인됐다. 색조(위 3개)는 이미 별도로
# 제외하니 여기엔 안 넣는다.
COSMETIC_ALLOWED_CATEGORIES = {
    "120000012",  # 스킨케어
    "120000017",  # UV케어
    "120000018",  # 바디・핸드・풋케어
    "120000019",  # 제모
    "120000020",  # 헤어
    "120000021",  # 네일
    "120000022",  # 향수
    "120000023",  # 맨즈뷰티
}
# [10으로 완화] 이 값은 두 군데에 쓰인다: (1) 검색결과에서 탐색대상 샵을
# 고를 때 (2) 크롤한 상품을 채택할 때. 상품 쪽 효과는 작지만(실측 +107건),
# 샵 쪽 효과가 크다 — 리뷰 6~9인 샵을 여태 통째로 무시하고 있었고, 그게
# 키워드 큐 고갈의 직접 원인이었다.
REVIEW_THRESHOLD = 10  # [v1.9.1부터 미사용] 상점선별 필터 해지됨. 과거 기록용으로만 남김
MIN_PRICE_JPY = 1500  # [v1.9.0부터 미사용] 상품저장 가격필터 해지됨. 과거 기록용으로만 남김
PRODUCT_SAVE_REVIEW_THRESHOLD = 10  # 20 -> 10 하향(품질 우선). 실측: 통합본 4,586건 중 리뷰<10이 4,355건(95%)이라 물량 손실은 5%뿐

STOPWORDS = ["選べる", "NEW", "セット", "公式", "限定", "特価", "お得", r"全\d+種", r"\bor\b", "×"]


def extract_core_keyword(title: str) -> str:
    t = title
    t = re.sub(r"[【\[（(][^】\])）]*[】\])）]", " ", t)
    t = re.split(r"\s*/", t)[0]
    t = re.sub(r"\d+\s*[枚mMｍＭlLｌＬgGｇＧ個点セ回本日%]+", " ", t)
    for sw in STOPWORDS:
        t = re.sub(sw, " ", t)
    t = re.sub(r"[,、]+\s*", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


VOLUME_RE = re.compile(r"[\d.]+\s*(?:ml|g|枚|個|本)(?:[×xX+][\d.]+\s*(?:ml|g|個|箱|セット))*")
BRACKET_RE = re.compile(r"[【\[（(][^】\])）]*[】\])）]")

# 1단계: 브랜드 보호 — 번역기가 브랜드명을 엉뚱하게 오역하는 걸 막기 위해
# 번역 전에 플레이스홀더로 바꿔치기하고, 번역 후 원래 브랜드로 복원한다.
KNOWN_BRANDS = [
    "AOU", "VT", "d'Alba", "TIRTIR", "SK-II", "rom&nd", "fwee", "hince",
    "dasique", "KAHI", "ATRUE", "AGAIN ME",
]

# 2단계: 화장품 특수 용어(의성어/의태어 등, 일반 번역기가 못 알아듣는 것들)
# — 실측으로 확인된 것만 우선 등록. 번역 전에 플레이스홀더로 바꿔치기하고
# 번역 후 화장품 업계 정식 표기로 복원한다(예: "ぽよん" → 일반번역은 "포옹"/
# "포용"이 되지만 화장품 정식표기는 "뽀용").
COSMETIC_TERM_MAP = {
    "ぽよん": "뽀용",
}


def _protect_and_translate(text: str, translator) -> str:
    """브랜드/특수용어를 플레이스홀더로 보호한 뒤 번역하고, 번역 후 복원한다."""
    protected = text
    restore_map = {}

    for i, brand in enumerate(KNOWN_BRANDS):
        if brand in protected:
            placeholder = f"XBRAND{i}X"
            protected = protected.replace(brand, placeholder)
            restore_map[placeholder] = brand

    for i, (jp_term, kr_term) in enumerate(COSMETIC_TERM_MAP.items()):
        if jp_term in protected:
            placeholder = f"XTERM{i}X"
            protected = protected.replace(jp_term, placeholder)
            restore_map[placeholder] = kr_term

    translated = translator.translate(protected) or protected

    for placeholder, original in restore_map.items():
        # 번역기가 플레이스홀더 대소문자/띄어쓰기를 살짝 바꿀 수 있어서 느슨하게 매칭
        translated = re.sub(re.escape(placeholder), original, translated, flags=re.IGNORECASE)
        translated = re.sub(placeholder.replace("X", r"X\s*"), original, translated, flags=re.IGNORECASE)

    return translated


SKIP_LOG_PATH = OUTPUT_DIR / "discovery_skip_log.json"
SEED_LOG_PATH = OUTPUT_DIR / "discovery_seed_log.json"
# [단순화] 아카이빙 로직은 이제 여기(1단계)에 없다 — 1단계는 순수하게
# 발굴+병합만 한다. "번역완료된 것을 아카이브로 옮기는" 책임은 2단계
# (auto_translate)로 옮겼다: 2단계가 번역을 마친 직후, discovery-live의
# discovery_state.json에서 방금 번역한 것들을 archive/로 옮긴다. 이렇게
# 하면 1단계가 다른 브랜치(translate-live)의 상태를 알 필요가 아예
# 없어져서 훨씬 단순해지고, 지금까지 겪었던 여러 버그(subprocess git lock
# 경합, 브랜치 기본값 불일치 등)의 근본 원인 자체가 사라진다.



def _append_skip_log(entries: list[dict]):
    """스킵 사유를 파일에 계속 누적 저장한다(나중에 '왜 27건만 남았나' 같은
    질문에 바로 답할 수 있도록)."""
    existing = []
    if SKIP_LOG_PATH.exists():
        existing = json.loads(SKIP_LOG_PATH.read_text(encoding="utf-8"))
    existing.extend(entries)
    SKIP_LOG_PATH.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")


def _append_seed_log(entries: list[dict]):
    """어떤 상품(베스트5 중 하나)에서 어떤 검색어(시드)가 나왔는지 전부
    기록한다 — 필터 통과여부와 무관하게 시드는 항상 생성되므로, 이 로그를
    보면 "왜 상점 발굴이 이렇게 뻗어나갔는지/막혔는지"를 바로 추적할 수 있다."""
    existing = []
    if SEED_LOG_PATH.exists():
        existing = json.loads(SEED_LOG_PATH.read_text(encoding="utf-8"))
    existing.extend(entries)
    SEED_LOG_PATH.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")


def crawl_shop_best5(shop_id: str) -> tuple[list[dict], bool]:
    """상점 베스트5를 전부 크롤링해서 (상품목록, 실패여부)를 반환한다.

    [9번 수정: 크롤실패 vs 무상품 구분] 예전엔 랭킹 조회가 실패해도, 상점에
    진짜 상품이 없어도 똑같이 빈 리스트를 돌려줬다. 그러면 호출하는 쪽이
    실패한 상점도 '방문완료(=상품없음)'로 영구 박제해버려서, 재시도할
    기회가 없어지고 시드도 안 생겨 큐 고갈로 이어졌다. 이제 실패
    (failed=True)와 진짜 무상품(failed=False, items=[])을 구분해서
    돌려주고, 호출하는 쪽(run())이 실패한 상점만 재시도하게 한다.

    필터 통과여부와 무관하게 5개 다 반환한다 — 사용자 지적사항 반영: 필터는
    "최종 결과물에 넣을지"만 결정해야지 "다음 검색 시드로 쓸지"까지 막으면
    안 된다. 색조/옵션이라도 다음 라운드 검색어로는 계속 써야 상점 발굴이
    5개씩 계속 뻗어나간다.

    각 item에 passes_filter(bool)와 skip_reason을 달아서 반환하고, 최종
    상품목록에 넣을지는 호출하는 쪽(run())이 passes_filter를 보고 결정한다.

    [주의] 필터링 단계(fetch_item_detail)와 번역 단계(GoogleTranslateSession)를
    반드시 분리된 두 단계로 처리해야 한다 — 둘 다 각자 sync_playwright()를
    쓰는데, 한쪽 세션이 열려있는 동안 다른 쪽을 또 열면 asyncio 충돌이 난다
    (앞서 multi_source_finder.py에서도 같은 문제를 겪었다)."""
    try:
        ranking = fetch_shop_ranking(shop_id)
    except ShopCrawlFailed as e:
        print(f"  [크롤실패-재시도대상] {shop_id}: {e}", file=sys.stderr)
        return [], True
    except Exception as e:  # noqa: BLE001 - 예상 못한 오류도 안전하게 실패로 본다
        print(f"  [크롤실패-예상외오류-재시도대상] {shop_id}: {type(e).__name__}: {e}", file=sys.stderr)
        return [], True
    if not ranking:
        return [], False  # 페이지는 열렸는데 진짜로 상품이 없음(정상 종결)

    all_items = []
    skip_entries = []
    for item in ranking:
        try:
            detail = fetch_item_detail(item["goods_no"], save_hires_image=False)
        except Exception as e:  # noqa: BLE001
            skip_entries.append({"shop_id": shop_id, "goods_no": item["goods_no"], "title": item["title"], "reason": f"상세조회실패: {e}"})
            continue

        category = detail.get("category_gdlc_cd")
        has_options = detail.get("has_options")
        # review_count가 None인 건 에러가 아니라 "리뷰가 아예 없어서 JSON-LD의
        # aggregateRating 필드 자체가 안 나오는" 정상 상태다 → 0으로 취급한다
        review_count = detail.get("review_count")
        if review_count is None:
            review_count = 0

        item["shop_id"] = shop_id
        item["category_gdlc_cd"] = category
        item["has_options"] = has_options
        item["review_count"] = review_count

        skip_reason = None
        price_jpy = item.get("price_jpy")
        if category in COLOR_COSMETIC_CATEGORIES:
            skip_reason = "색조카테고리"
        elif category not in COSMETIC_ALLOWED_CATEGORIES:
            skip_reason = "화장품카테고리아님"
        elif has_options:
            skip_reason = "옵션있음"
        elif review_count >= PRODUCT_SAVE_REVIEW_THRESHOLD:
            skip_reason = f"리뷰수{review_count}(20개이상)"
        # [v1.9.0에서 완전해지 -> 방금 재조정] 가격 필터(1500엔 이하)는
        # 계속 해지 상태로 둔다. 리뷰수만 20개 미만으로 다시 걸어달라는
        # 요청 반영 — PRODUCT_SAVE_REVIEW_THRESHOLD=20 (구
        # REVIEW_THRESHOLD=10보다 완화된 값. 상점선별 단계는 여전히
        # v1.9.1대로 리뷰수 무관하게 전부 방문후보로 삼는다 — 이건
        # '어떤 상점에 들어갈지'가 아니라 '그 상점에서 크롤한 상품 중
        # 뭘 최종목록에 저장할지'만 다시 제한하는 것이다.

        item["passes_filter"] = skip_reason is None
        item["skip_reason"] = skip_reason

        if skip_reason:
            print(f"    [필터탈락-{skip_reason}] {item['goods_no']} {item['title'][:30]} (그래도 다음 시드로는 사용)", file=sys.stderr)
            skip_entries.append({"shop_id": shop_id, "goods_no": item["goods_no"], "title": item["title"], "reason": skip_reason, "category": category})
        else:
            print(f"    [저장] {item['goods_no']} review={review_count} {item['title'][:30]}", file=sys.stderr)

        all_items.append(item)  # 필터 통과여부와 무관하게 항상 추가(시드 생성용)

    if skip_entries:
        _append_skip_log(skip_entries)

    return all_items, False


def find_low_review_shops(keyword: str, visited_shops: set) -> list[dict]:
    """[해지] 예전엔 검색결과 중 review_count<REVIEW_THRESHOLD인 것만
    방문후보로 삼았다. 사용자 지시로 이 필터도 없앤다 — 이제 검색결과에
    나온 모든 상점이 방문후보가 된다(이미 방문한 상점만 제외). 실측
    확인(검증기록 2번): 리뷰필터 있으면 24개, 없으면 28개로 큰 차이가
    아니었지만, 저장단계 필터 해지(v1.9.0)와 함께 상점 선별 단계까지
    완전히 열어서 "리뷰수와 무관하게 전부 발굴"로 방향을 통일한다."""
    html = search_qoo10(keyword)
    results = parse_results(html)
    seen = {}
    for r in results:
        if r["shop_id"] not in visited_shops:
            seen[r["shop_id"]] = r
    return list(seen.values())


def _state_path(suffix: str | None = None) -> "Path":
    if suffix:
        return OUTPUT_DIR / f"discovery_state_{suffix}.json"
    return STATE_PATH


def load_peer_visited(my_suffix: str | None) -> set:
    """다른 워커들의 discovery_state_*.json에서 visited_shops만 읽어온다.

    [설계원칙: 파일 소유자 1명] 각 워커는 자기 파일에만 쓰고, 남의 파일은
    읽기만 한다. 그래서 공유 파일에 여러 명이 쓰다 깨지는 사고(과거 실패
    #2 index.lock 경합, #4 push 실패로 상점 128개 유실)가 구조적으로
    일어날 수 없다.

    이게 없으면 워커마다 자기 방문기록만 봐서 남이 이미 판 샵을 또 판다 —
    실측 중복률 샵 84.3%, 키워드 76.8%였다.
    """
    peers = set()
    for p in sorted(OUTPUT_DIR.glob("discovery_state_*.json")):
        if my_suffix is not None and p.name == f"discovery_state_{my_suffix}.json":
            continue
        try:
            peers |= set(json.loads(p.read_text(encoding="utf-8")).get("visited_shops", []))
        except Exception:  # noqa: BLE001
            continue  # 남이 쓰는 중이라 깨져 보일 수 있다 — 그냥 건너뛴다
    return peers


def _load_state(suffix: str | None = None) -> dict:
    path = _state_path(suffix)
    if path.exists():
        state = json.loads(path.read_text(encoding="utf-8"))
        state.setdefault("failed_shops", {})  # 하위호환: 예전 상태파일엔 없음
        return state
    return {"visited_shops": [], "all_products": [], "shop_urls": [], "pending_keywords": None, "seen_keywords": [], "failed_shops": {}}


def _save_state(state: dict, suffix: str | None = None):
    path = _state_path(suffix)
    path.parent.mkdir(exist_ok=True, parents=True)
    path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def crawl_shop_best5_with_timeout(shop_id: str, timeout_seconds: int = 90) -> tuple[list[dict], bool]:
    """crawl_shop_best5를 별도 프로세스로 격리해서 실행하고, timeout_seconds를
    넘기면 그 프로세스를 강제 종료(SIGKILL)한다.

    [중요: signal.alarm에서 subprocess 격리 방식으로 교체함] 예전엔
    같은 프로세스 안에서 signal.alarm(90)으로 시간제한을 걸었는데,
    실측으로 이게 신뢰할 수 없다는 게 확인됐다 — Playwright의 페이지
    로딩 예외처리 도중 signal이 전달되지 않고 워커 하나가 3시간 넘게
    완전히 멈춰버린 사고가 있었다(원인: 자식 스레드/이벤트루프 안에서
    발생한 예외 처리 중에는 파이썬 시그널 핸들러가 실행을 못 미룰 수
    있음). subprocess.run(timeout=N)은 그 자식 프로세스 내부에서 무슨
    일이 일어나든 운영체제 수준에서 확실하게 죽이므로 100% 신뢰할 수
    있다.

    [9번 수정] 타임아웃/비정상종료/파싱실패도 전부 '크롤실패'로 본다 —
    (items=[], failed=True)를 돌려주면 run()이 이 상점을 방문완료로
    박제하지 않고 다음 사이클에 다시 시도한다."""
    try:
        result = subprocess.run(
            [sys.executable, "crawl_single_shop.py", shop_id],
            capture_output=True,
            text=True,
            timeout=timeout_seconds,
            cwd=str(Path(__file__).resolve().parent),
        )
    except subprocess.TimeoutExpired:
        print(f"  [TIMEOUT-재시도대상] 상점 {shop_id} 처리가 {timeout_seconds}초를 넘어 강제종료, 다음으로 넘어감")
        return [], True

    if result.returncode != 0:
        print(f"  [ERROR-재시도대상] 상점 {shop_id} 서브프로세스 비정상종료(code={result.returncode}): {result.stderr[-500:] if result.stderr else ''}")
        return [], True

    try:
        payload = json.loads(result.stdout.strip() or '{"items": [], "failed": true}')
    except json.JSONDecodeError:
        print(f"  [ERROR-재시도대상] 상점 {shop_id} 결과 파싱 실패: {result.stdout[-500:]}")
        return [], True

    return payload.get("items", []), bool(payload.get("failed", False))



def run(keyword_ja: str, target_products: int, max_shops: int | None = None, shops_per_keyword: int | None = None, seed_keywords: list[str] | None = None, state_suffix: str | None = None):
    state = _load_state(state_suffix)
    visited_shops = set(state["visited_shops"])
    all_products = {p["goods_no"]: p for p in state["all_products"]}
    shop_urls = state["shop_urls"]
    if state["pending_keywords"] is not None:
        pending_keywords = state["pending_keywords"]
    elif seed_keywords:
        pending_keywords = list(seed_keywords)
    else:
        pending_keywords = [keyword_ja]
    seen_keywords = set(state["seen_keywords"])
    failed_shops = dict(state.get("failed_shops", {}))  # {shop_id: 실패누적횟수}
    MAX_CRAWL_RETRIES = 3  # 이만큼 실패하면 포기하고 방문완료로 넘긴다(영구루프 방지)

    if state["visited_shops"]:
        print(f"[RESUME] 상점 {len(visited_shops)}개, 상품 {len(all_products)}건부터 이어서 진행")

    def save():
        nonlocal all_products
        _save_state(
            {
                "visited_shops": list(visited_shops),
                "all_products": list(all_products.values()),
                "shop_urls": shop_urls,
                "pending_keywords": pending_keywords,
                "seen_keywords": list(seen_keywords),
                "keyword_stats": keyword_stats,
                "failed_shops": failed_shops,
            },
            state_suffix,
        )

    # [v4.2.0] 검색어 길이별 성과 측정.
    #  생성되는 검색어의 23.2%가 단어 하나짜리(パウダー 등)인데, 이게
    #  유리한지(넓게 걸려 다양한 상점 노출) 불리한지(대형 상점만 떠서
    #  한 사이클 낭비) 측정된 적이 없다. 발굴은 상점 1곳당 검색어 5.4개를
    #  쓴다 — 5개 중 4개는 새 상점을 못 만든다는 뜻이고, 그 낭비가 어디서
    #  오는지 알아야 큐 순서를 손댈지 판단할 수 있다.
    #
    #  파일이 커지지 않도록 개별 로그가 아니라 '토큰 수 구간별 누적'만
    #  담는다(구간 4개 x 숫자 3개).
    keyword_stats = state.get("keyword_stats") or {}

    def _kw_bucket(text: str) -> str:
        n = len(text.split())
        return "1" if n <= 1 else ("2" if n == 2 else ("3-4" if n <= 4 else "5+"))

    def _record_kw(text: str, shops_found: int, saved: int) -> None:
        b = _kw_bucket(text)
        cur = keyword_stats.setdefault(b, {"keywords": 0, "shops": 0, "saved": 0})
        cur["keywords"] += 1
        cur["shops"] += shops_found
        cur["saved"] += saved

    while pending_keywords and len(all_products) < target_products:
        if max_shops and len(visited_shops) >= max_shops:
            print(f"\n[STOP] 최대 상점수({max_shops}) 도달")
            break

        kw = pending_keywords[0]  # 상점 처리 끝나야 pop (중간에 끊겨도 재개 가능)
        if kw in seen_keywords:
            pending_keywords.pop(0)
            save()
            continue

        print(f"\n[검색] {kw}")
        products_before = len(all_products)
        # 내 방문기록 + 다른 워커 방문기록을 합쳐서 제외한다(중복 크롤 방지).
        peer_visited = load_peer_visited(state_suffix)
        shops = find_low_review_shops(kw, visited_shops | peer_visited)
        if peer_visited:
            print(f"  (다른 워커 방문분 {len(peer_visited)}곳 제외)")
        if shops_per_keyword:
            shops = shops[:shops_per_keyword]
        print(f"  -> 신규 저리뷰 상점 {len(shops)}개 (이 검색어에서 처리할 상점)")

        # [중대버그 수정] 예전엔 캡(STEP=3)에 걸려 중단돼도 아래에서 이 검색어를
        # "다 봤다"고 표시하고 버렸다. 검색어 하나가 샵 30곳을 물어와도 3곳만
        # 받고 27곳을 영구 폐기한 셈이라, 검색어 재생산율이 1 밑으로 떨어져
        # 큐가 지수적으로 말라붙었다(실측 재생산율 0.58). 이제 캡 때문에
        # 끊겼으면 검색어를 큐에 그대로 남겨 다음 호출에서 이어받는다.
        # 이미 방문한 샵은 find_low_review_shops가 걸러주므로 중복도 없다.
        interrupted_by_cap = False
        for shop in shops:
            if max_shops and len(visited_shops) >= max_shops:
                interrupted_by_cap = True
                break
            if len(all_products) >= target_products:
                break
            shop_id = shop["shop_id"]
            shop_urls.append(f"https://m.qoo10.jp/shop/{shop_id}")
            print(f"\n  [상점진입] {shop_id} (review={shop['review_count']}, 실패이력={failed_shops.get(shop_id, 0)}회)")

            crawled_items, crawl_failed = crawl_shop_best5_with_timeout(shop_id)

            if crawl_failed:
                # [9번 수정] 크롤 자체가 실패한 상점은 '방문완료'로 박제하지
                # 않는다 — 예전엔 실패든 진짜 무상품이든 똑같이 방문완료로
                # 표시해버려서, 실패한 상점에서 나왔어야 할 시드가 영원히
                # 안 생기고 조용히 큐가 말라붙었다. 실패 횟수를 세어
                # MAX_CRAWL_RETRIES(3회)까지는 재시도 대상으로 남기고,
                # 그 이상 계속 실패하면(진짜 고장난 상점일 수 있음) 그때
                # 포기하고 방문완료 처리해 무한루프를 막는다.
                failed_shops[shop_id] = failed_shops.get(shop_id, 0) + 1
                if failed_shops[shop_id] >= MAX_CRAWL_RETRIES:
                    print(f"  [포기] {shop_id} {MAX_CRAWL_RETRIES}회 연속 크롤실패 — 방문완료 처리하고 넘어감")
                    visited_shops.add(shop_id)
                else:
                    print(f"  [재시도예정] {shop_id} 크롤실패({failed_shops[shop_id]}/{MAX_CRAWL_RETRIES}회) — 방문완료 처리 안 함")
                save()
                continue

            visited_shops.add(shop_id)
            failed_shops.pop(shop_id, None)  # 이번엔 성공했으니 실패 이력 초기화
            seed_entries = []
            for item in crawled_items:
                # 시드는 필터 통과여부와 무관하게 전부 생성(사용자 지적사항 반영)
                core = extract_core_keyword(item["title"])
                if core:
                    pending_keywords.append(core)
                    seed_entries.append(
                        {
                            "from_shop": shop_id,
                            "from_goods_no": item["goods_no"],
                            "from_title": item["title"],
                            "passes_filter": item.get("passes_filter"),
                            "seed_keyword": core,
                        }
                    )
                # 최종 상품목록에는 필터 통과한 것만 넣는다
                if item.get("passes_filter") and len(all_products) < target_products:
                    all_products[item["goods_no"]] = item
            if seed_entries:
                _append_seed_log(seed_entries)
            save()  # 매 상점마다 저장 (타임아웃 걸려도 이어서 진행 가능)

        if interrupted_by_cap:
            print(f"  [보류] 캡 도달로 중단 — 검색어를 큐에 남겨둠(다음 호출에서 이어서): {kw[:40]}")
            save()
            break

        _record_kw(kw, len(shops), len(all_products) - products_before)
        seen_keywords.add(kw)
        pending_keywords.pop(0)
        save()

    print(f"\n[DONE] 상점 {len(visited_shops)}개 방문, 상품 {len(all_products)}건 확보")
    return list(all_products.values()), shop_urls


def export_excel(products: list[dict], out_path: str):
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = [
        {
            "큐텐상품번호": p.get("goods_no"),
            "상점ID": p.get("shop_id"),
            "상품명(원본)": p.get("title"),
            "한글검증명칭": p.get("name_kr_verified"),
            "브랜드": p.get("brand"),
            "가격(엔)": p.get("price_jpy"),
            "리뷰수": p.get("review_count"),
            "옵션있음": p.get("has_options"),
            "카테고리코드": p.get("category_gdlc_cd"),
            "상품URL": p.get("item_url"),
        }
        for p in products
    ]
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="저리뷰상품", index=False)

    from openpyxl import load_workbook

    wb = load_workbook(out_path)
    ws = wb.active
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    widths = {"A": 14, "B": 16, "C": 45, "D": 40, "E": 16, "F": 10, "G": 8, "H": 10, "I": 14, "J": 45}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(out_path)
    print(f"[EXCEL] {out_path} 저장 완료 ({len(rows)}행)")


def main():
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)
    keyword_ja = sys.argv[1]
    target = int(sys.argv[2])
    out_path = sys.argv[3]
    max_shops = int(sys.argv[4]) if len(sys.argv) > 4 else None
    shops_per_keyword = int(sys.argv[5]) if len(sys.argv) > 5 and sys.argv[5].strip() else None
    state_suffix = sys.argv[6] if len(sys.argv) > 6 and sys.argv[6].strip() else None

    before = 0
    state_file = _state_path(state_suffix)
    if state_file.exists():
        try:
            before = len(json.loads(state_file.read_text(encoding="utf-8")).get("all_products", []))
        except Exception:  # noqa: BLE001
            before = 0

    products, shop_urls = run(keyword_ja, target, max_shops, shops_per_keyword, state_suffix=state_suffix)

    # [커밋폭탄 차단] 엑셀은 ZIP 컨테이너라 내용이 같아도 생성시각이 박혀
    # 매번 다른 파일이 된다. 예전엔 수확이 0건이어도 무조건 다시 써서,
    # 워크플로의 "변경 있을 때만 커밋" 방어가 무력화됐다 — 워커 12개가
    # 2초에 한 번씩 빈 커밋을 찍어 브랜치가 86,269커밋까지 불어났다.
    # 이제 새로 얻은 게 없으면 엑셀을 건드리지 않는다.
    if len(products) > before or not Path(out_path).exists():
        export_excel(products, out_path)
    else:
        print(f"[SKIP] 신규 수확 0건 — 엑셀 재생성 생략(불필요한 커밋 방지, 누적 {len(products)}건)")
    print(f"\n방문한 상점: 누적 {len(shop_urls)}개 (최근 5개: {shop_urls[-5:]})")


if __name__ == "__main__":
    main()