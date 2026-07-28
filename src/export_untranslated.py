"""
export_untranslated.py — 통합본에서 아직 번역 안 된 상품을 엑셀로 뽑는다.

[왜 필요한가] Claude API 자동번역을 파이프라인에서 전부 걷어냈다(사용자
방침: 번역은 별도 Claude 창에서 직접 처리). 그래서 파이프라인은
"수확 -> 발굴 -> 통합"까지만 하고, 미번역 상품을 엑셀로 내보낸다.
사용자가 그 엑셀의 '한글 상품명(번역)' 열을 채워서 돌려주면
import_translated.py로 다시 통합본에 반영한다.

사용법:
    python export_untranslated.py <state.json> <out.xlsx> [최대건수]
"""

import json
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

CATEGORY_NAMES = {
    "120000012": "스킨케어",
    "120000017": "UV케어",
    "120000018": "바디/핸드/풋",
    "120000019": "제모",
    "120000020": "헤어",
    "120000013": "색조",
    "120000014": "색조",
    "120000016": "색조",
}
HANGUL_RE = re.compile(r"[가-힣]")


def export(state_path: str, out_path: str, limit: int | None = None) -> int:
    path = Path(state_path)
    if not path.exists():
        print(f"[SKIP] {path} 없음")
        return 0

    state = json.loads(path.read_text(encoding="utf-8"))
    products = state.get("all_products", [])
    pending = [p for p in products if not p.get("translated_kr")]
    if limit:
        pending = pending[:limit]

    print(f"[INFO] 전체 {len(products)}건 중 미번역 {len(pending)}건")
    if not pending:
        print("[INFO] 미번역 상품 없음 — 엑셀 생성 생략")
        return 0

    rows = []
    for i, p in enumerate(pending, 1):
        title = p.get("title") or ""
        rows.append({
            "번호": i,
            "큐텐상품번호": p.get("goods_no"),
            "큐텐 상품명(원문)": title,
            "한글 상품명(번역)": "",  # 여기를 채워서 돌려주면 된다
            "한글 포함 여부": "한글 있음" if HANGUL_RE.search(title) else "일본어만",
            "브랜드": p.get("brand") or "",
            "가격(엔)": p.get("price_jpy"),
            "리뷰수": p.get("review_count"),
            "카테고리": CATEGORY_NAMES.get(p.get("category_gdlc_cd"), p.get("category_gdlc_cd") or ""),
            "상점ID": p.get("shop_id") or "",
            "상품URL": p.get("item_url") or "",
        })

    df = pd.DataFrame(rows)
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="미번역상품", index=False)

    wb = load_workbook(out)
    ws = wb.active
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", start_color="C0392B", end_color="C0392B")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # 번역 입력칸(D열)은 노란 헤더로 눈에 띄게
    ws["D1"].fill = PatternFill("solid", start_color="E8B923", end_color="E8B923")
    ws["D1"].font = Font(name="Arial", bold=True, color="000000", size=11)
    ws.row_dimensions[1].height = 28

    widths = {"A": 6, "B": 14, "C": 58, "D": 44, "E": 14, "F": 18,
              "G": 11, "H": 8, "I": 12, "J": 16, "K": 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    fill_in = PatternFill("solid", start_color="FFF9E0", end_color="FFF9E0")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=cell.column_letter in ("C", "D"))
        ws.row_dimensions[row[0].row].height = 32
        r = row[0].row
        ws[f"D{r}"].fill = fill_in
        ws[f"G{r}"].number_format = "#,##0"
        for col in ("A", "E", "G", "H"):
            ws[f"{col}{r}"].alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(out)

    print(f"[DONE] {len(rows)}건 -> {out}")
    return len(rows)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    limit = int(sys.argv[3]) if len(sys.argv) > 3 and sys.argv[3].strip() else None
    export(sys.argv[1], sys.argv[2], limit)
